#!/usr/bin/env python3
# =============================================================================
# NEXUS PI DISASSEMBLER (BytesOfPI.xlsm)
# -----------------------------------------------------------------------------
# Purpose:
#   Treat the workbook as a receiver: extract the header-byte stream (PI ASM Analysis),
#   infer a minimal "opcode" trace that explains how each header could arise from
#   recent prior headers via a small verb set (unary + binary transforms),
#   and write a disassembly report + plots.
#
# Nexus lens:
#   - headers are "state"
#   - transitions are "verbs"
#   - bytes are "exhaust"/projection
# =============================================================================

import os, math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

# ----------------------------- helpers ----------------------------------------

def nibble_pair(byte: int):
    """Return (hi, lo) nibbles (0..15) from a byte."""
    return (byte >> 4) & 0xF, byte & 0xF

def pack_pair(a: int, b: int) -> int:
    """Pack (hi, lo) nibbles into a byte."""
    return ((a & 0xF) << 4) | (b & 0xF)

# --------------------------- disassembler core --------------------------------

def explain_sequence_window(bytes_list, window=16):
    """
    Greedy, local disassembler:
      For each target byte at time t, try to explain it as:
        - unary transform of any prior byte within window
        - binary transform of any two prior bytes within window
      Choose the closest (lowest time-distance) explanation.

    Opcodes (verbs):
      Unary:
        HOLD:          x -> x
        SUM_mod16:     (a,b) -> (a, (a+b) mod 16)
        DIFFSUM_mod16: (a,b) -> (|b-a| mod 16, (a+b) mod 16)
        SUM:           (a,b) -> (a, a+b)      [only if a+b <= 15]
        DIFFSUM:       (a,b) -> (|b-a|, a+b)  [only if a+b <= 15]
      Binary:
        XOR2:          x,y -> x xor y
        ADD2:          x,y -> (x+y) mod 256
        DIFF2:         x,y -> |x-y| mod 256

    Returns:
      DataFrame with inferred opcode + source indices.
    """
    hist = []
    expl = []

    for t, byte in enumerate(bytes_list):
        hi, lo = nibble_pair(byte)
        best = None  # tuple(score, opcode, detail, src1, src2)

        # unary on any prior within window
        for j in range(max(0, len(hist) - window), len(hist)):
            prev = hist[j]
            a, b = nibble_pair(prev)
            candidates = [
                ("HOLD", prev, f"from t{j}: {prev:02X}"),
                ("SUM_mod16", pack_pair(a, (a + b) & 0xF),
                 f"from t{j}: ({a:X},({a:X}+{b:X}) mod16)"),
                ("DIFFSUM_mod16", pack_pair(abs(b - a) & 0xF, (a + b) & 0xF),
                 f"from t{j}: (|{b:X}-{a:X}|,(+ ) mod16)"),
            ]
            s = a + b
            if s <= 15:
                candidates += [
                    ("SUM", pack_pair(a, s), f"from t{j}: ({a:X},{a:X}+{b:X})"),
                    ("DIFFSUM", pack_pair(abs(b - a), s),
                     f"from t{j}: (|{b:X}-{a:X}|,{a:X}+{b:X})"),
                ]

            for name, val, desc in candidates:
                if val == byte:
                    dist = t - j
                    score = dist  # prefer closest
                    candidate = (score, name, desc, j, None)
                    if best is None or candidate[0] < best[0]:
                        best = candidate

        # binary on any pair within window
        for j in range(max(0, len(hist) - window), len(hist)):
            for k in range(j + 1, len(hist)):
                x = hist[k]
                y = hist[j]
                candidates = [
                    ("XOR2", x ^ y, f"from t{k} xor t{j}: {x:02X} xor {y:02X}"),
                    ("ADD2", (x + y) & 0xFF, f"from t{k} + t{j}: ({x:02X}+{y:02X}) mod256"),
                    ("DIFF2", (abs(x - y)) & 0xFF, f"from t{k} - t{j}: |{x:02X}-{y:02X}|"),
                ]
                for name, val, desc in candidates:
                    if val == byte:
                        # slightly penalize binary so unary explanations win if equally close
                        dist = max(t - k, t - j)
                        score = dist + 0.5
                        candidate = (score, name, desc, k, j)
                        if best is None or candidate[0] < best[0]:
                            best = candidate

        if best:
            _, opcode, detail, src1, src2 = best
            matched = True
        else:
            opcode, detail, src1, src2, matched = "PUSH", "", None, None, False

        hist.append(byte)
        expl.append((t, byte, f"{byte:02X}", hi, lo, opcode, matched, src1, src2, detail))

    return pd.DataFrame(
        expl,
        columns=["t", "byte", "hex", "hi", "lo", "opcode", "matched", "src1", "src2", "detail"]
    )

# ------------------------------- runner ---------------------------------------

def run(xlsm_path: str, outdir: str = "pi_disasm_out", window: int = 16, max_rows: int = 5000):
    os.makedirs(outdir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)
    if "PI ASM Analysis" not in wb.sheetnames:
        raise ValueError("Workbook missing sheet 'PI ASM Analysis'.")

    ws = wb["PI ASM Analysis"]

    # Read (index, header, data_byte) rows
    rows = []
    for r in range(2, max_rows + 1):
        idx = ws.cell(r, 1).value
        if idx is None:
            continue
        header = ws.cell(r, 2).value
        data_byte = ws.cell(r, 3).value
        rows.append((idx, str(header).zfill(2), data_byte))

    df = pd.DataFrame(rows, columns=["index", "header_hex", "data_byte"])
    # parse headers as bytes
    bytes_list = [int(h, 16) for h in df["header_hex"].tolist()]

    dis = explain_sequence_window(bytes_list, window=window)
    dis["index"] = df["index"].values
    dis["data_byte"] = df["data_byte"].values

    # Save outputs
    dis.to_csv(os.path.join(outdir, "pi_header_disassembly.csv"), index=False)

    op_counts = dis["opcode"].value_counts().reset_index()
    op_counts.columns = ["opcode", "count"]
    op_counts.to_csv(os.path.join(outdir, "opcode_counts.csv"), index=False)

    # Recurrence (repeat gaps)
    last = {}
    reps = []
    for t, hx in enumerate(dis["hex"].tolist()):
        if hx in last:
            reps.append((t, hx, t - last[hx]))
        last[hx] = t
    rep_df = pd.DataFrame(reps, columns=["t", "hex", "gap"])
    rep_df.to_csv(os.path.join(outdir, "header_repeats.csv"), index=False)

    # Plots
    plt.figure(figsize=(10, 4))
    plt.plot(dis["t"], dis["byte"])
    plt.xlabel("t (byte index)")
    plt.ylabel("Header byte value")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(outdir, "header_byte_trace.png"), dpi=160)
    plt.close()

    plt.figure(figsize=(9, 4))
    plt.bar(op_counts["opcode"], op_counts["count"])
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("Count")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(outdir, "opcode_counts.png"), dpi=160)
    plt.close()

    plt.figure(figsize=(10, 3))
    plt.plot(dis["t"], dis["matched"].astype(int))
    plt.ylim(-0.1, 1.1)
    plt.xlabel("t")
    plt.ylabel("Matched (1=yes)")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(outdir, "match_over_time.png"), dpi=160)
    plt.close()

    if len(rep_df) > 0:
        plt.figure(figsize=(8, 4))
        plt.hist(rep_df["gap"], bins=30)
        plt.xlabel("Repeat gap (steps)")
        plt.ylabel("Count")
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(os.path.join(outdir, "repeat_gap_hist.png"), dpi=160)
        plt.close()

    # Summary
    summary = []
    summary.append(f"Rows disassembled: {len(dis)}")
    summary.append(f"Matched: {dis['matched'].mean()*100:.1f}%")
    summary.append("Opcode counts:")
    for _, row in op_counts.iterrows():
        summary.append(f"  {row['opcode']}: {row['count']}")
    summary.append(f"Unique headers: {len(set(dis['hex'].tolist()))}")
    summary.append(f"Repeat events: {len(rep_df)}")
    if len(rep_df) > 0:
        summary.append(f"Median repeat gap: {rep_df['gap'].median():.1f}")
        summary.append(f"Most common gap: {rep_df['gap'].value_counts().idxmax()}")

    with open(os.path.join(outdir, "summary.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(summary))

    print("DONE. Outputs in:", outdir)

if __name__ == "__main__":
    run("BytesOfPI.xlsm", outdir="pi_disasm_out", window=16)
